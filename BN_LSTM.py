
from openpyxl import Workbook
from openpyxl import load_workbook
import tensorflow as tf
from random import choice, shuffle
import numpy as np
from numpy import array
import matplotlib.pyplot as plt
import sys
import bn_lstmCell

data_file_name='filterResult.xlsx'
data_sheet_name='Sheet7'	#预测和训练的时候切换到不同的excel表单
data_row_name='A'

time_step=40      #time_step
rnn_unit=200      #hidden layer units, assume each layer has the same unit
batch_size=60     #batch size
input_size=1      #dim of input
output_size=1     #dim of output
lr=0.0006         #learning rate
is_train= 1 #0 for test, 1 for train

layer_number=1	  #the LSTM layer number
is_drop=0         #0 for not dropping, 1 for dropping
drop_out_rate = 1  #rate of not dropping out

####################################################################
#load data from excel
#return data as (-1,1)
wb = load_workbook(data_file_name)
ws_old = wb.get_sheet_by_name(data_sheet_name)
list1 = []
new_predict = []
for row in ws_old.iter_rows(data_row_name):
    try:
        if row[0].value is not None:
          x = float(row[0].value)
    except ValueError:
        continue
    if x != 0:
        list1.append(x)
data=array(list1,ndmin=2).transpose()
#######################################################################
#whiten the data
#produce the data as [batch,time_step,dim_of_input]
normalize_data = (data-np.mean(data))/np.std(data)

train_x,train_y=[],[]
for i in range(len(normalize_data)-time_step-1):
    x=normalize_data[i:i+time_step]
    y=normalize_data[i+1:i+time_step+1]
    train_x.append(x.tolist())
    train_y.append(y.tolist())


######################################################################
#public tensor 

X=tf.placeholder(tf.float32, [None,time_step,input_size])
Y=tf.placeholder(tf.float32, [None,time_step,output_size])

#weights and biases of input layer and output layer
weights={
     'in':tf.Variable(tf.random_normal([input_size,rnn_unit])),
     'out':tf.Variable(tf.random_normal([rnn_unit,1]))
     }
biases={
    'in':tf.Variable(tf.constant(0.1,shape=[rnn_unit,])),
    'out':tf.Variable(tf.constant(0.1,shape=[1,]))
    }
######################################################################
def model(batch):
    #LSTM model part
    ################################
    w_in=weights['in']
    b_in=biases['in']
    input=tf.reshape(X,[-1,input_size])  						#turn the tensor into 2 dimension to calculate
    input_rnn=tf.matmul(input,w_in)+b_in 						#y=wx+b
    input_rnn=tf.reshape(input_rnn,[-1,time_step,rnn_unit])  	#turn back to the 3 dimension as the input of cell

    if layer_number == 1:									 	#if only one layer, then single cell
    	cell=bn_lstmCell.BN_LSTMCell(rnn_unit,is_train)
        #cell=tf.contrib.rnn.BasicLSTMCell(rnn_unit)
        #cell=bn_lstmCell.BN_LSTMCell()
		#tf.contrib.rnn.BasicLSTMCell(rnn_unit)
    elif is_drop==0:										 	#if multiple layers without dropping out, then use MultiRNNCell
    	cell=tf.contrib.rnn.MultiRNNCell([create_cell() for i in range(layer_number)])
    else:													 	#if multiple layers with a dropping out layer, then add the dr
    	list1=[dropout_cell()]
    	list1.extend([create_cell() for i in range(layer_number-1)])#only add dropout layer in the first layer output
    	cell=tf.contrib.rnn.MultiRNNCell(list1)

    init_state = cell.zero_state(batch,dtype=tf.float32)
    output_rnn,final_states=tf.nn.dynamic_rnn(cell, input_rnn,initial_state=init_state, dtype=tf.float32)  #output_rnn是记录lstm每个输出节点的结果，final_states是最后一个cell的结果

    output=tf.reshape(output_rnn,[-1,rnn_unit])            		#turn the tensor into 2 dimension to calculate
    w_out=weights['out']			
    b_out=biases['out']
    pred=tf.matmul(output,w_out)+b_out
    return pred,final_states
    ################################

def create_cell():
    cell=tf.contrib.rnn.BasicLSTMCell(rnn_unit)
    return cell

def dropout_cell():
    cell=tf.contrib.rnn.BasicLSTMCell(rnn_unit)
    lstm_cell = tf.contrib.rnn.DropoutWrapper(cell,output_keep_prob=drop_out_rate)
    return lstm_cell

def train():
    global batch_size
    pred,_=model(batch_size)
    #loss function, square loss
    loss=tf.reduce_mean(tf.square(tf.reshape(pred,[-1])-tf.reshape(Y, [-1])))
    train_op=tf.train.AdamOptimizer(lr).minimize(loss)
    saver=tf.train.Saver(tf.global_variables())

    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        #train 10000 times
        for i in range(10000):
            step=0
            start=0
            end=start+batch_size
            while(end<len(train_x)):
                _,loss_=sess.run([train_op,loss],feed_dict={X:train_x[start:end],Y:train_y[start:end]})
                start+=batch_size
                end=start+batch_size
                #每10步保存一次参数
                if step%10==0:
                    print(i,step,loss_)    
                step+=1
            print("model save：",saver.save(sess,'result3\\layer_'+str(layer_number)+'_isdropr_'+str(is_drop)+'_rate_'+str(drop_out_rate)+'.model'))

def predict():
    #tf.get_variable_scope().reuse_variables()
    pred,_=model(2)    #input[1,time_step,input_size]
    saver=tf.train.Saver(tf.global_variables())
    sess = tf.Session()
	#parameter fetch
    sess.run(tf.global_variables_initializer())
    module_file = tf.train.latest_checkpoint('result3')
    saver.restore(sess, module_file) 
    prev_seq=train_x[0]
    predict=[]
    for i in range(len(train_x)-2):
    	prev_seq = train_x[i]
    	next_seq=sess.run(pred,feed_dict={X:train_x[i:i+2]})
    	predict.extend(next_seq[-1])
	#print(predict)
    global new_predict
    for i in range(len(predict)):
      new_predict.append(predict[i]*np.std(data)+np.mean(data))
    new_predict=new_predict[:-50]
    print(accuracy(0.1))
    print(accuracy(0.01))
    print(accuracy(0.001))
    plt.figure()
    plt.plot(list(range(len(normalize_data))), normalize_data, color='b')
    #plt.plot(list(range(len(normalize_data), len(normalize_data) + len(predict))), predict, color='r')
    plt.plot(list(range(time_step-1,time_step-1+len(predict))), predict, color='r')
    plt.savefig('bnlayer_'+str(layer_number)+'_isdropr_'+str(is_drop)+'_rate_'+str(drop_out_rate)+'.png')
    plt.show()
    

def accuracy(confidence=0):
	count = 0
	for i in range(len(new_predict)):
	      if abs(new_predict[i]-data[i+time_step+1])<=confidence:
	          count = count+1
	return count/27660

def main():
	if float(is_train)==0:
		drop_out_rate=1
		predict()
	else:
		train()
predict()





